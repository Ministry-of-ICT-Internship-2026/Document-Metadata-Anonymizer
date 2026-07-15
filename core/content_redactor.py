from __future__ import annotations

import math
import os
import re
import zipfile
from collections import Counter
from typing import Optional

from lxml import etree

try:
    import fitz
except ImportError:
    fitz = None

PATTERNS = {
    "email": re.compile(
        r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b'
    ),
    "phone": re.compile(
        r'\b(?:\+?\d{1,3}[-.\s]?)?\(?\d{2,4}\)?[-.\s]?\d{3,4}[-.\s]?\d{3,5}\b'
    ),
    "nin": re.compile(
        r'\b\d{2}[-]\d{6}[-]\d{6,7}[-]\d{2}\b'
    ),
    "credit_card": re.compile(
        r'\b(?:\d[ -]*?){13,19}\b'
    ),
    "ipv4": re.compile(
        r'\b(?:(?:25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(?:25[0-5]|2[0-4]\d|[01]?\d\d?)\b'
    ),
    "ipv6": re.compile(
        r'\b(?:'
        r'(?:[A-Fa-f0-9]{1,4}:){7}[A-Fa-f0-9]{1,4}'
        r'|(?:[A-Fa-f0-9]{1,4}:){1,7}:'
        r'|(?:[A-Fa-f0-9]{1,4}:){1,6}:[A-Fa-f0-9]{1,4}'
        r'|(?:[A-Fa-f0-9]{1,4}:){1,5}(?::[A-Fa-f0-9]{1,4}){1,2}'
        r'|(?:[A-Fa-f0-9]{1,4}:){1,4}(?::[A-Fa-f0-9]{1,4}){1,3}'
        r'|(?:[A-Fa-f0-9]{1,4}:){1,3}(?::[A-Fa-f0-9]{1,4}){1,4}'
        r'|(?:[A-Fa-f0-9]{1,4}:){1,2}(?::[A-Fa-f0-9]{1,4}){1,5}'
        r'|[A-Fa-f0-9]{1,4}:(?::[A-Fa-f0-9]{1,4}){1,6}'
        r'|:(?::[A-Fa-f0-9]{1,4}){1,7}'
        r'|::(?:[A-Fa-f0-9]{1,4}:){0,5}[A-Fa-f0-9]{1,4}'
        r'|[Ff][Ee]80:(?::[A-Fa-f0-9]{1,4}){0,4}%[0-9a-zA-Z]{1,}'
        r')\b'
    ),
    "jwt": re.compile(
        r'\beyJ[A-Za-z0-9_-]+\.[A-Za-z0-9_-]+\.[A-Za-z0-9_-]+\b'
    ),
    "api_key": re.compile(
        r'\b(?:sk-|pk-|ghp_|gho_|ghu_|ghs_|ghr_|AKIA|AKI)[A-Za-z0-9_-]{8,}\b'
    ),
    "ref_code": re.compile(
        r'\b[A-Z]{2,}[-/][A-Za-z0-9]{2,}[-/]\d{4}[-/]\d{2,}\b'
    ),
    "currency": re.compile(
        r'\b(?:UGX|USD|EUR|KES|GBP|JPY|CNY|ZAR|NGN|TZS|RWF|BIF|SSP|CDF|ETB'
        r'|AUD|CAD|CHF|INR|SAR|AED|MXN|BRL|NGN|GHS|ZMW|MZN|AOA)'
        r'\s*[\d,]+(?:\.\d{1,2})?\b'
    ),
}

REDACTED_LABEL = "[REDACTED]"

CONFIDENTIAL_KEYWORDS = [
    "confidential", "secret", "classified",
    "internal only", "internal review", "internal use only",
    "not for distribution", "not for circulation",
    "not for external circulation", "not for external distribution",
    "private", "proprietary",
    "password", "passwd", "pwd",
    "bank account", "account number", "routing number",
    "salary", "compensation", "bonus",
    "ssn", "social security", "passport",
    "national id", "driver license", "drivers license",
    "patient", "medical record", "diagnosis",
    "trade secret", "nda", "non-disclosure",
    "access code", "pin", "security code",
    "cvv", "cvc",
    "username", "login",
    "date of birth", "dob",
    "top secret", "restricted",
    "cabinet", "budget allocation",
    "internal briefing", "internal memo",
]

WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
DOCX_TEXT_PARTS = [
    "word/document.xml",
    "word/header1.xml",
    "word/header2.xml",
    "word/header3.xml",
    "word/footer1.xml",
    "word/footer2.xml",
    "word/footer3.xml",
]

_keyword_re_cache: dict[str, re.Pattern] = {}
_sentence_re = re.compile(r'[^.!?\n]*[.!?\n]')


def _build_keyword_re(keywords: list[str]) -> re.Pattern:
    key = "|".join(sorted(keywords))
    if key not in _keyword_re_cache:
        _keyword_re_cache[key] = re.compile(
            r'\b(' + '|'.join(re.escape(k) for k in keywords) + r')\b',
            re.IGNORECASE,
        )
    return _keyword_re_cache[key]


def shannon_entropy(s: str) -> float:
    if not s or not s.strip():
        return 0.0
    s = s.strip()
    counts = Counter(s)
    total = len(s)
    return -sum((c / total) * math.log2(c / total) for c in counts.values())


def luhn_check(s: str) -> bool:
    digits = [int(c) for c in s if c.isdigit()]
    if len(digits) < 13 or len(digits) > 19:
        return False
    checksum = 0
    for i, d in enumerate(reversed(digits)):
        if i % 2 == 1:
            d *= 2
            if d > 9:
                d -= 9
        checksum += d
    return checksum % 10 == 0


def _find_matches(text: str, rules: set[str],
                  custom_keywords: list[str] | None = None) -> list[tuple[int, int, str]]:
    matches: list[tuple[int, int, str]] = []

    if "email" in rules and "email" in PATTERNS:
        for m in PATTERNS["email"].finditer(text):
            if m.start() == m.end():
                continue
            matches.append((m.start(), m.end(), "email"))

    if "phone" in rules and "phone" in PATTERNS:
        for m in PATTERNS["phone"].finditer(text):
            if m.start() == m.end():
                continue
            matches.append((m.start(), m.end(), "phone"))

    if "nin" in rules and "nin" in PATTERNS:
        for m in PATTERNS["nin"].finditer(text):
            matches.append((m.start(), m.end(), "nin"))

    if "credit_card" in rules and "credit_card" in PATTERNS:
        for m in PATTERNS["credit_card"].finditer(text):
            raw = m.group()
            if luhn_check(raw):
                matches.append((m.start(), m.end(), "credit_card"))

    if "ip" in rules:
        if "ipv4" in PATTERNS:
            for m in PATTERNS["ipv4"].finditer(text):
                if m.start() == m.end():
                    continue
                matches.append((m.start(), m.end(), "ip"))
        if "ipv6" in PATTERNS:
            for m in PATTERNS["ipv6"].finditer(text):
                if m.start() == m.end():
                    continue
                matches.append((m.start(), m.end(), "ip"))

    if "jwt" in rules and "jwt" in PATTERNS:
        for m in PATTERNS["jwt"].finditer(text):
            matches.append((m.start(), m.end(), "jwt"))

    if "api_key" in rules and "api_key" in PATTERNS:
        for m in PATTERNS["api_key"].finditer(text):
            matches.append((m.start(), m.end(), "api_key"))

    if "ref_code" in rules and "ref_code" in PATTERNS:
        for m in PATTERNS["ref_code"].finditer(text):
            matches.append((m.start(), m.end(), "ref_code"))

    if "currency" in rules and "currency" in PATTERNS:
        for m in PATTERNS["currency"].finditer(text):
            matches.append((m.start(), m.end(), "currency"))

    if "high_entropy" in rules:
        for m in re.finditer(r'\S{12,}', text):
            tok = m.group()
            if len(tok) >= 12 and shannon_entropy(tok) > 4.0:
                matches.append((m.start(), m.end(), "high_entropy"))

    if "confidential" in rules:
        kw = custom_keywords or CONFIDENTIAL_KEYWORDS
        kw_re = _build_keyword_re(kw)
        for sent_m in _sentence_re.finditer(text):
            sent = sent_m.group()
            if kw_re.search(sent):
                matches.append((sent_m.start(), sent_m.end(), "confidential"))

    matches.sort(key=lambda x: (x[0], -x[1]))
    filtered: list[tuple[int, int, str]] = []
    for s, e, t in matches:
        if not filtered or s >= filtered[-1][1]:
            filtered.append((s, e, t))
    return filtered


def redact_text(text: str, rules: set[str],
                custom_keywords: list[str] | None = None) -> tuple[str, dict[str, int]]:
    if not text:
        return text, {}
    matches = _find_matches(text, rules, custom_keywords)
    if not matches:
        return text, {}

    counts: dict[str, int] = {}
    result = []
    pos = 0
    for s, e, t in matches:
        result.append(text[pos:s])
        result.append(REDACTED_LABEL)
        counts[t] = counts.get(t, 0) + 1
        pos = e
    result.append(text[pos:])
    return "".join(result), counts


def _merge_counts(a: dict[str, int], b: dict[str, int]) -> dict[str, int]:
    merged = dict(a)
    for k, v in b.items():
        merged[k] = merged.get(k, 0) + v
    return merged


def _redact_docx(filepath: str, output_path: str,
                 rules: set[str], custom_keywords: list[str] | None) -> dict[str, int]:
    total: dict[str, int] = {}

    with zipfile.ZipFile(filepath, "r") as zin:
        infos = zin.infolist()
        parts: dict[str, bytes] = {}

        for info in infos:
            fname = info.filename
            data = zin.read(fname)

            if fname in DOCX_TEXT_PARTS or fname.startswith("word/charts/"):
                pass
            elif re.search(r'word/(header|footer)\d*\.xml$', fname):
                pass
            else:
                parts[fname] = data
                continue

            try:
                root = etree.fromstring(data)
            except etree.XMLSyntaxError:
                parts[fname] = data
                continue

            changed = False
            ns_map = {"w": WORD_NS}
            for elem in root.iter(f"{{{WORD_NS}}}t"):
                if elem.text:
                    new_text, counts = redact_text(elem.text, rules, custom_keywords)
                    if new_text != elem.text:
                        elem.text = new_text
                        total = _merge_counts(total, counts)
                        changed = True
            if changed:
                parts[fname] = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
            else:
                parts[fname] = data

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for info in infos:
                data = parts.get(info.filename)
                if data is None:
                    data = zin.read(info.filename)
                zout.writestr(info, data)

    return total


def _redact_xlsx(filepath: str, output_path: str,
                 rules: set[str], custom_keywords: list[str] | None) -> dict[str, int]:
    import openpyxl
    total: dict[str, int] = {}
    wb = openpyxl.load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    new_text, counts = redact_text(cell.value, rules, custom_keywords)
                    if new_text != cell.value:
                        cell.value = new_text
                        total = _merge_counts(total, counts)

    wb.save(output_path)
    wb.close()
    return total


def _redact_pdf(filepath: str, output_path: str,
                rules: set[str], custom_keywords: list[str] | None) -> dict[str, int]:
    if fitz is None:
        raise RuntimeError("PyMuPDF (fitz) is not available — cannot redact PDF content")

    doc = fitz.open(filepath)
    total: dict[str, int] = {}

    for page_num, page in enumerate(doc):
        text = page.get_text("text")
        if not text.strip():
            continue
        new_text, page_counts = redact_text(text, rules, custom_keywords)
        if new_text == text:
            continue
        total = _merge_counts(total, page_counts)

        for s, e, _t in _find_matches(text, rules, custom_keywords):
            match_str = text[s:e]
            if not match_str.strip():
                continue
            try:
                rects = page.search_for(match_str)
                for rect in rects:
                    page.add_redact_annot(rect, fill=(0, 0, 0))
            except Exception:
                pass

        page.apply_redactions()

    doc.save(output_path, deflate=True, garbage=4)
    doc.close()
    return total


def redact_content(filepath: str, output_path: str,
                   filters: dict | None = None) -> dict[str, int] | None:
    if not filters:
        return None

    rules: set[str] = set()
    rule_map = {
        "redact_email": "email",
        "redact_phone": "phone",
        "redact_nin": "nin",
        "redact_credit_card": "credit_card",
        "redact_ip": "ip",
        "redact_api_key": "api_key",
        "redact_jwt": "jwt",
        "redact_confidential": "confidential",
        "redact_high_entropy": "high_entropy",
        "redact_ref_code": "ref_code",
        "redact_currency": "currency",
    }
    for filter_key, rule_name in rule_map.items():
        if filters.get(filter_key, False):
            rules.add(rule_name)

    if not rules:
        return None

    custom_keywords: list[str] | None = None
    raw = filters.get("redact_custom_keywords", "")
    if isinstance(raw, str) and raw.strip():
        custom_keywords = [kw.strip() for kw in raw.split(",") if kw.strip()]

    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".docx":
        return _redact_docx(filepath, output_path, rules, custom_keywords)
    elif ext == ".xlsx":
        return _redact_xlsx(filepath, output_path, rules, custom_keywords)
    elif ext == ".pdf":
        return _redact_pdf(filepath, output_path, rules, custom_keywords)
    return None
