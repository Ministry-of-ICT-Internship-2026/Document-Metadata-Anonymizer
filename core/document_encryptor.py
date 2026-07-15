from __future__ import annotations

import base64
import os
import re
import zipfile

from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from lxml import etree

try:
    import fitz
except ImportError:
    fitz = None

ENC_PREFIX = "⧙ENC:"
ENC_SUFFIX = ":ENC⧘"
WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
DOCX_TEXT_PARTS = [
    "word/document.xml",
    "word/header1.xml",
    "word/header2.xml",
    "word/header3.xml",
    "word/footer1.xml",
    "word/footer2.xml",
    "word/footer3.xml",
]


_global_salt: bytes | None = None


def _derive_key(password: str, salt: bytes) -> bytes:
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=200000)
    return base64.urlsafe_b64encode(kdf.derive(password.encode()))


def _fernet_from_password(password: str) -> tuple[Fernet, bytes]:
    salt = os.urandom(16)
    key = _derive_key(password, salt)
    return Fernet(key), salt


def _fernet_from_password_with_salt(password: str, salt: bytes) -> Fernet:
    key = _derive_key(password, salt)
    return Fernet(key)


def _get_fernet(password: str) -> Fernet:
    global _global_salt
    if _global_salt is None:
        _global_salt = os.urandom(16)
    return _fernet_from_password_with_salt(password, _global_salt)


def encrypt_text(text: str, fernet: Fernet) -> str:
    if not text or not text.strip():
        return text
    if text.startswith(ENC_PREFIX):
        return text
    token = fernet.encrypt(text.encode()).decode()
    return f"{ENC_PREFIX}{token}{ENC_SUFFIX}"


def decrypt_text(text: str, fernet: Fernet) -> str:
    if not text or not text.strip():
        return text
    pattern = re.escape(ENC_PREFIX) + r"(.+?)" + re.escape(ENC_SUFFIX)
    m = re.search(pattern, text)
    if not m:
        return text
    try:
        return fernet.decrypt(m.group(1).encode()).decode()
    except Exception:
        return text


def _encrypt_docx(filepath: str, output_path: str, password: str) -> bytes:
    fernet, salt = _fernet_from_password(password)

    with zipfile.ZipFile(filepath, "r") as zin:
        infos = zin.infolist()
        parts: dict[str, bytes | None] = {}

        for info in infos:
            fname = info.filename
            data = zin.read(fname)

            should_process = (
                fname in DOCX_TEXT_PARTS
                or bool(re.search(r'word/(header|footer)\d*\.xml$', fname))
            )
            if not should_process:
                parts[fname] = data
                continue

            try:
                root = etree.fromstring(data)
            except etree.XMLSyntaxError:
                parts[fname] = data
                continue

            for elem in root.iter(f"{{{WORD_NS}}}t"):
                if elem.text:
                    elem.text = encrypt_text(elem.text, fernet)

            parts[fname] = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
            # Write salt marker first
            salt_info = zipfile.ZipInfo("encryption_salt.bin")
            zout.writestr(salt_info, salt)
            for info in infos:
                data = parts.get(info.filename)
                if data is None:
                    data = zin.read(info.filename)
                zout.writestr(info, data)

    return salt


def _decrypt_docx(filepath: str, output_path: str, password: str) -> None:
    salt = _read_salt(filepath)
    if salt is None:
        raise ValueError("Not an encrypted document (no salt found)")
    fernet = _fernet_from_password_with_salt(password, salt)

    with zipfile.ZipFile(filepath, "r") as zin:
        infos = [i for i in zin.infolist() if i.filename != "encryption_salt.bin"]
        parts: dict[str, bytes | None] = {}

        for info in infos:
            fname = info.filename
            data = zin.read(fname)

            should_process = (
                fname in DOCX_TEXT_PARTS
                or bool(re.search(r'word/(header|footer)\d*\.xml$', fname))
            )
            if not should_process:
                parts[fname] = data
                continue

            try:
                root = etree.fromstring(data)
            except etree.XMLSyntaxError:
                parts[fname] = data
                continue

            for elem in root.iter(f"{{{WORD_NS}}}t"):
                if elem.text:
                    elem.text = decrypt_text(elem.text, fernet)

            parts[fname] = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for info in infos:
                data = parts.get(info.filename)
                if data is None:
                    data = zin.read(info.filename)
                zout.writestr(info, data)


def _encrypt_xlsx(filepath: str, output_path: str, password: str) -> bytes:
    import openpyxl
    fernet, salt = _fernet_from_password(password)
    wb = openpyxl.load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = encrypt_text(cell.value, fernet)

    wb.save(output_path)
    wb.close()

    with zipfile.ZipFile(output_path, "a") as zout:
        zout.writestr("encryption_salt.bin", salt)

    return salt


def _decrypt_xlsx(filepath: str, output_path: str, password: str) -> None:
    import openpyxl
    salt = _read_salt(filepath)
    if salt is None:
        raise ValueError("Not an encrypted document (no salt found)")
    fernet = _fernet_from_password_with_salt(password, salt)
    wb = openpyxl.load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = decrypt_text(cell.value, fernet)

    wb.save(output_path)
    wb.close()


def _read_salt(filepath: str) -> bytes | None:
    if not os.path.isfile(filepath) or filepath.endswith(".pdf"):
        return None
    try:
        with zipfile.ZipFile(filepath, "r") as z:
            return z.read("encryption_salt.bin")
    except (KeyError, zipfile.BadZipFile):
        return None


def is_encrypted(filepath: str) -> bool:
    return _read_salt(filepath) is not None


def encrypt_document(filepath: str, output_path: str, password: str) -> None:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".docx":
        _encrypt_docx(filepath, output_path, password)
    elif ext == ".xlsx":
        _encrypt_xlsx(filepath, output_path, password)
    elif ext == ".pdf":
        raise ValueError("PDF text encryption is not supported — content streams cannot be reversibly encrypted while keeping the file openable.")
    else:
        raise ValueError(f"Unsupported format: {ext}")


def decrypt_document(filepath: str, output_path: str, password: str) -> None:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".docx":
        _decrypt_docx(filepath, output_path, password)
    elif ext == ".xlsx":
        _decrypt_xlsx(filepath, output_path, password)
    elif ext == ".pdf":
        raise ValueError("PDF text decryption is not supported.")
    else:
        raise ValueError(f"Unsupported format: {ext}")
