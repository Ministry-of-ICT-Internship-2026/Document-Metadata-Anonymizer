import os
import sys

import pytest
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from core.xlsx_stripper import XlsxMetadataError, read_metadata, strip_metadata

SAMPLES = os.path.join(os.path.dirname(__file__), "sample_files")


def _s(name):
    p = os.path.join(SAMPLES, name)
    if not os.path.exists(p):
        pytest.skip(f"{name} not found")
    return p


def test_read_metadata():
    m = read_metadata(_s("plain_metadata.xlsx"))
    assert m["creator"] == "John Doe"
    assert m["lastModifiedBy"] == "jdoe_mdaict"
    assert m["title"] == "Draft Spreadsheet"


def test_strip_clears_all(tmp_path):
    src = _s("plain_metadata.xlsx")
    out = str(tmp_path / "stripped.xlsx")
    before = strip_metadata(src, out)
    assert before["creator"] == "John Doe"
    after = read_metadata(out)
    assert after["creator"] == ""


def test_hidden_sheets_revealed(tmp_path):
    src = _s("plain_metadata.xlsx")
    out = str(tmp_path / "stripped.xlsx")
    b = read_metadata(src)
    assert "HiddenSheet" in b["hidden_sheets"]
    strip_metadata(src, out)
    wb = openpyxl.load_workbook(out)
    assert all(ws.sheet_state == "visible" for ws in wb.worksheets)
    wb.close()


def test_content_preserved(tmp_path):
    src = _s("plain_metadata.xlsx")
    out = str(tmp_path / "stripped.xlsx")
    wb = openpyxl.load_workbook(src)
    v = wb.active["A1"].value
    wb.close()
    strip_metadata(src, out)
    wb2 = openpyxl.load_workbook(out)
    assert wb2.active["A1"].value == v
    wb2.close()


def test_no_metadata(tmp_path):
    m = read_metadata(_s("no_metadata.xlsx"))
    strip_metadata(_s("no_metadata.xlsx"), str(tmp_path / "out.xlsx"))
    assert os.path.exists(str(tmp_path / "out.xlsx"))


def test_missing_raises():
    with pytest.raises(XlsxMetadataError):
        read_metadata("/tmp/nonexistent.xlsx")
